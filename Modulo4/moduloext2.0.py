#####################
def cls():
    ##limpando terminal
    from os import system, name
    if name == 'posix':
        return system('clear')
    return system('cls')
# Continuação...

# Web Scraping com Python usando requests e bs4 BeautifulSoup
# Web Scraping é o ato de "raspar a web" buscando informações de forma
# automatizada, com determinada linguagem de programação, para uso posterior.
# O módulo requests consegue carregar dados da Internet para dentro do seu
# código. Já o bs4.BeautifulSoup é responsável por interpretar os dados HTML
# em formato de objetos Python para facilitar a vida do desenvolvedor.
# Doc: https://www.crummy.com/software/BeautifulSoup/bs4/doc.ptbr/
# Instalação
# pip install requests types-requests bs4

import requests
from bs4 import BeautifulSoup

url = 'https://books.toscrape.com/'
response = requests.get(url) #busco o html
raw_html = response.text #faço o retorno
#e aqui eu já começo a fazer a raspagem
parsed_html = BeautifulSoup(raw_html, 'html.parser')

# caso você quiser o título, basta executar essa funcao
def exemplo():
    # tô verificando se é None porque pode retornar duas coisas, sendo: None e o valor.
    if parsed_html.title is not None:
        #e aqui eu pego o título do texto da página
        print(parsed_html.title.text)


selecao_qualquer = parsed_html.select_one('#default > div > div > div > div > div.page-header.action > h1')
# print(selecao_qualquer)
#se eu quiser saber quem é a tag mãe dele:
print(selecao_qualquer.parent) # é o div a tag mãe.

####################Explicação de termo
# Dangling comma	(vírgula pendurada) ou Trailling comma (virgula final) nada mais é que uma prática realizada
# para deixar uma vírgula após o último elemento de uma lista, tupla, etc... EX:  
v0 = (1,2,3,4,)
v1 = [0,1,2,]
#como pode ver, funcionou normalmente.
# Dizem que deixa o código mais organizado e facilita na hora da leitura,
# Na hora de duplicar, ajuda também. 
print(v0,'\n',v1)

###########################################################################
## Selenium
# https://selenium-python.readthedocs.io/locating-elements.html
# Primeiro, vou precisar instalar um driver para uso. Como utilizarei o compre, com base nos requesitos do meu sistema,
# devo fazer a procura no seguinte site: https://developer.chrome.com/docs/chromedriver/downloads
# Agora, preciso instalar o selenium com o seguinte cmd:
# pip install selenium
def selenium1(): #coloquei dentro de uma função para não automatizar após a execução
    #agora, vamos chamar o módulo selenium
    from selenium import webdriver
    # Além do selenium, vamos precisar do service, que é um outro módulo complementar do selenium
    from selenium.webdriver.chrome.service import Service

    #agora, vamos precisar definir o caminho da pasta atual. Para isso, precisaremos do módulo path inicialmente
    from pathlib import Path
    #agora, vamos definir esse diretório
    PASTA_ATUAL = Path(__file__).parent
    #agora, vamos criar uma para o chromedrive
    CHROMEDRIVER = PASTA_ATUAL / 'M4Selenium' / 'chromedriver'

    #agora, vamos definir algumas opções.
    # Quando queremos abrir o chrome com algumas opções a mais ou a menos, fazemos uso de uma variável 
    # que guardará ChromeOptions. Essas opções faz parte do que você quer que o nevegador faça quando abrir. Por exemplo,
    # se você passa --headless, você não verá a interface do navegador. Se houvesse --disable-gpu, a gpu não seria utilizada durante o processo.
    #veja mais flags para options em: https://peter.sh/experiments/chromium-command-line-switches/
    chrome_opcoes = webdriver.ChromeOptions()
    #para indicar o serviço (driver) que estamos fazendo uso:
    chrome_servico = Service(executable_path=CHROMEDRIVER)
    #e para isso funcionar, precisamos passar todos os parâmetros anteriores em um lugar só:
    chrome_navegador = webdriver.Chrome(
        service=chrome_servico,
        options=chrome_opcoes,
    )

    #agora, para fazer uma simples demonstração, vou abrir um site qualquer
    chrome_navegador.get('https://youtube.com') #como pode ver, o site abre e após milésimos, ele fecha.
    #se fosse passado algo em options, como '--headless', a interface do navegador não iria abrir.
cls()
###########

def SeleniumWW():
    #mesma estrutura anterior
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By #esse módulo vai ser útil para selecionar um elemento dentro da página
    from selenium.webdriver.support.wait import WebDriverWait #Como os elementos não aparecem de imediato, vamos precisar de algo que vai fazer com que o código dê um tempo de espera para o elemento aparecer.
    from selenium.webdriver.support import expected_conditions as EC

    #definimos o caminho do webdrive
    from pathlib import Path
    PASTA_ATUAL = Path(__file__).parent
    CHROMEDRIVER = PASTA_ATUAL / 'M4Selenium' / 'chromedriver'

    #agora os parâmetros
    chrome_opcoes = webdriver.ChromeOptions()
    chrome_servico = Service(executable_path=CHROMEDRIVER)
    chrome_navegador = webdriver.Chrome(
        service=chrome_servico,
        options=chrome_opcoes,
    )

    #e o comando abaixo, para abrir o navegador na url seguinte:
    chrome_navegador.get('https://www.youtube.com/')

    #por fim, vamos iniciar a automação com uma busca simples por tags de nome, que no caso abaixo, será para
    # um input padrão lá na barra de pesquisa do yt
    from time import sleep
    procurar_input = WebDriverWait(chrome_navegador, 10).until( #aguarda 10 segundos para dar tempo para a tag aparecer
        EC.presence_of_element_located(
            (By.NAME, 'search_query') #busca por uma tag name 
        )
    )
    #a tag search_query é um input, ou seja, após localizada, vamos escrever alguma coisa.
    procurar_input.send_keys('hello world!!!!!!!!!!!!!!!!!!') #esse comando pode ser utilizado para escrever algo atraés da tag
    sleep(3) #e após escrever, ele aguarda 3 segundos. 

    ####################
    #Mas e se eu quisesse fazer uso de alguma tecla do teclado?
    from selenium.webdriver.common.keys import Keys
    #agora, após aquele hello world, ele vai apertar enter.
    procurar_input.send_keys(Keys.ENTER)
    sleep(3) #apenas para que a execução não acabe de imediato

    #agora, vou fazer uma coleta dos resultados que me ocorreram
    resultado_pesquisas = chrome_navegador.find_element(By.ID, 'contents')

    #agora, vou filtrar apenas pelo elemento que guarda o link (normalmente é a tag "a")
    links = resultado_pesquisas.find_elements(By.TAG_NAME, "a")
    links[0].click()
    sleep(3)

    #voltar aqui no futuro, pois pode ser que a função de efetuar o click em vídeo acabe se tornando obsoleta!

    '''
    Com o Selenium aprendido, vou deixá-lo dentro de uma função, para que não fique executando e atrapalhando
    o restante do código.
    '''

#############################################################################################
# Resuminho do Luiz:
# 
# Usando subprocess para executar e comandos externos
# subprocess é um módulo do Python para executar
# processos e comandos externos no seu programa.
# O método mais simples para atingir o objetivo é usando subprocess.run().
# Argumentos principais de subprocess.run():
# - stdout, stdin e stderr -> Redirecionam saída, entrada e erros
# - capture_output -> captura a saída e erro para uso posterior
# - text -> Se True, entradas e saídas serão tratadas como texto
# e automaticamente codificadas ou decodificadas com o conjunto
# de caracteres padrão da plataforma (geralmente UTF-8).
# - shell -> Se True, terá acesso ao shell do sistema. Ao usar
# shell (True), recomendo enviar o comando e os argumentos juntos.
# - executable -> pode ser usado para especificar o caminho
# do executável que iniciará o subprocesso.
# Retorno:
# stdout, stderr, returncode e args
# Importante: a codificação de caracteres do Windows pode ser
# diferente. Tente usar cp1252, cp852, cp850 (ou outros). Linux e
# mac, use utf_8.
# Comando de exemplo:
# Windows: ping 127.0.0.1
# Linux/Mac: ping 127.0.0.1 -c 4

#importando módulo
import subprocess

#vamos listar um comando que ele vai executar (caso linux)
comando = ['ping', '127.0.0.1', '-c', '4']

#e agora, vamos executar o comando acima
processo = subprocess.run(
    comando,
    #esse cmd abaixo não tinha antes
    capture_output=True, #o intuito desse cmd é capturar e exibir a saída
    #aqui eu posso passar:
    # text=True -> ele iria pegar automaticamente a codificação do seu SO, evitando o uso do decode (No windows pode não funcionar bem)

    #e poderia passar também
    # encoding='utf_8' -> isso evitaria o uso dos comandos acima, já que a saída já iria utilizar, por padrão,o nome da codificação que vocẽ informar como argumento.
)

#agora, vamos fazer a visualização
print(processo) #aqui ele mostra todas as informações referente ao processo. 
# No entanto, caso eu queira filtrar, bastaria digitar o nome do atributo
print(processo.args) #para visualizar os argumentos
print(processo.stdout) #para visualizar a saída
print(processo.stderr) #mostra os erros erro (se não houver, o retorno será vazio)
print(processo.returncode) #retorna um bool em caso de ocorrência de erro

cls()
#vou colocar um argumento no parâmetro capture_output
# Agora, quando eu fizer uso do strout, ele vai retornar o texto formatado de forma estranha.
print(processo.stdout)

# É por isso que vamos ter que tratar, seja através a inserção de um segundo parâmetro, que é o caso do text=True
# Ou  fazemos uso do método decode, sendo:

cls()
print(processo.stdout.decode('utf_8'))#aqui você deve colocar o nome do encoding (codificação) que o seu SO utiliza
#no windows, a codificação utf8 pode não funcionar bem e com base nos testes do Luiz, o que mais funcionou foi a codificação cp850

#Caso você tente fazer alguma coisa através do módulo sobprocess e começar a dar pau, mesmo estando com a syntax correta, volte na aula 321

############### curiosidade à parte
'''Outra forma de descobrir qual SO estou usando, com base no kernel'''
import sys
print(sys.platform) #no meu caso, retornou linux


####################################################################################################################
# Aula 323 e 324
#Jupyter notebook -> uma ide, assim como o vs code, porém é mais adequado para cientistas de dados.
#é recomendado SEMPRE fazer a instalação do Jupyter em um ambiente virtual devido a possibilidade do uso de texto, markdown, html, etc.

# Comando de instalação:
# pip install notebook

#após a instalação, basta digitar o seguinte comando: 
#jupyter notebook
#após digitar esse comando, será aberto uma instância local da sua máquina na porta 8888. Esse é o Jupyter.

#após salvar um arquivo, vai perceber que ele foi salvo num formato ipynp. É do jupyter

#pesquisar sobre matplotlib, numpy + pandas e dataframe depois

#Esse tal de notebook realmente é bem útil quando se trata de execução e exemplificação de código.

#o próprio vscode consegue carregar notebooks.

""" 
Alguns comandos:

    # -> significa <h1> (texto de título maior no html)
    ## -> h2 
    ### ...
    <h1>conteudo</h1> -> html também funciona
    <code> print("olá mundo!")</code> -> para código
    '''print("opa")''' -> para codigo também.

Basicamente, você pode utilizar markdown e html para isso. 

"""

##############################################################
# Módulo Thread: Trabalhando com múltiplos processos
#Vamos supor que você precisa fazer duas ou mais execuções no código ao mesmo tempo, como faria isso?
# Thread é um módulo que permite fazer uso das threads do seu processador para executar mais de um processo, que no nosso caso,
# será um código
from time import sleep

#problema: quero executar uma contagem até 10 e quero que em paralelo, aconteça uma outra coisa sem interferir diretamente
#na contagem.
for __ in range(10):
    print(__)
    #eu poderia colocar a ação aqui, mas em alguns casos, isso iria interferir na contagem
    # sleep(1)

#Por isso, podemos utilizar um thread para dividir os processos.
from threading import Thread #importo a classe Thread

#uma maneira de criar um Thread é através de uma classe, aplicando o conceito de Herança em POO
class MeuThread(Thread):
    #inicializamos o nosso construtor e definimos os parâmetros que vamo precisar
    def __init__(self, nome, tempo):
        #definimos agora os atributos de instância
        self.nome = nome
        self.tempo = tempo
        #e por fim, inicializamos o construtor da classe Thread
        super().__init__()
    
    #na classe thread, há um método denominado run. Após aplicar a Herança, vamos aplicar o polimofirsmo para realizar
    #a ação que desejamos no código.
    def run(self):
        #no caso, eu quero que apareça o nome de uma pessoa durante aquele intervalo de 0 à 10 fazendo uso do multitarefa.
        # sleep(self.tempo) 
        print(self.nome)

#agora, vamos criar um objeto e passar os argumentos para a classe
t1 = MeuThread('jao', 5)
#e para inicializar a classe, fazemos uso do método start.
t1.start()

#por fim, repitimos o for. 
for __ in range(10):
    print(__)
    # sleep(1)
# e como pode ver, durante a contagem, ele dxecutará o print do nome do indivíduo na tela.
#Se quiser realizar o teste efetivo, basta tirar os comentários do sleep.

"""
######----------
# Outra forma de executar a thread:
def vai_demorar(texto, tempo):
    sleep(tempo)
    print(texto)

#agora, vamos chamar a thread
thread1 = Thread(target=vai_demorar, args=('Hello worlds', 2))
thread1.start()
#target é o nome da func que será executada
#args são os argumentos da função.
# Detalhe: Se a função tiver apenas um argumento, deverá utilizar o conceito de Dangling comma para prevenir bugs. 

#-Agora, vamos fazer um for que vai executar junto com a thread
for __ in range(10):
    print(__)
    sleep(.5)

#basicamente, é isso.
#mas e como eu sei se a thread ativou ou não?

thread2 = Thread(target=vai_demorar, args=('Aoba', 2))
thread2.start()
while thread2.is_alive(): #enquanto a thread não estivar, o código abaixo executará
    sleep(2)
    print('thread ainda não ativou')

""" #-> vou fazer uso de docs-strings para continuar usando este módulo.

#####Thread é um módulo muito interessante para trabalhar com multitarefas, mas também ter que saber trabalhar. 
#vou trazer um exemplo de um caso:

class Ingresso:
    #no nosso construtor, vamos colocar a quantidade de estoque que temos para vender
    def __init__(self, estoque):
        self.estoque = estoque
    
    #agora, quando alguém quiser comprar, precisaremos vender.
    def comprar(self, quantidade):
        self.quantidade = quantidade
    
        #agora, temos que ver se temos estoque antes de vender.
        if self.estoque < self.quantidade:
            print('não temos ingresso suficiente!')
            return
        
        sleep(1) #comente essa linha e o código vai normalizar. (logo vocẽ vai entender o porquê.)

        #se tivermos estoque, bora vender com base na quantidade!
        self.estoque -= quantidade

        #informa ao usuário a transação
        print(f'você comprou {self.quantidade} ingresso(s)!\nAinda temos {self.estoque} ingressos!')

#inicializamos o objeto
ingressos = Ingresso(estoque=20)

#e por fim, bora comprar com thread.
if __name__ == 'underfined': #se quiser testar, troque underfined para __main__
    for i in range(20):
        t = Thread(target=ingressos.comprar, args=(5,)).start() #aqui, ele vai comprar em loop com todas as threads
        #contanto que não haja um tempo de espera devido a uma consulta no banco de dados, requesição ou algo do tipo, vai 
        #funcionar perfeitamente.

        #no entanto, se houver alguma espera, o código gera transtornos silenciosos. Vamos simular isso colocando um sleep.
        #após colocar o sleep, percebeu que começou a vir números negativos? então. Isso é porque todas as threads passaram 
        #juntas pelo if.

        #Como todas passaram juntas pelo if e tiveram de esperar no sleep, as outras continuaram passando (pois não houve 
        #decremento)
        
        #isso é apenas uma simulação doq pode ocorrer em um banco de dados online real, tipo supabase.
        #por isso, temos que tratar e uma das soluções para isso é fazer uso do módulo lock, que seu obejtivo é guardar 
        #estados.'


'''
Eu vou repetir o mesmo código para exemplificar o uso do Lock. Sei que poderia utilizar herança + polimofrismo para esse 
exemplo, mas acho que ficaria um pouco complexo para visualizaar. Logo, vou reutilizar o mesmo modelo da classe Integresso.
'''
from threading import Thread, Lock #agora,vamos importar o Lock.
class Ingresso2: #criar um nome semelhante para a classe
    #no nosso construtor, vamos colocar a quantidade de estoque que temos para vender
    def __init__(self, estoque):
        self.estoque = estoque
        self.lock = Lock() #criamos uma instância para o lock (para facilitar a nossa vida)
    
    #agora, quando alguém quiser comprar, precisaremos vender.
    def comprar(self, quantidade):
        #agora,vamos inibir a passagem de mais de uma thread ao mesmo tempo
        
        self.lock.acquire() #com isso, quando uma thread passar por aqui, as outras ficarão aqui aguardando.
        
        self.quantidade = quantidade 
        #agora, temos que ver se temos estoque antes de vender.
        if self.estoque < self.quantidade:
            print('não temos ingresso suficiente!')
            #e quando a quantidade de ingressos se esgostarem, devemos deixar outra thread passar
            self.lock.release()
            return
        
        sleep(1)

        #se tivermos estoque, bora vender com base na quantidade!
        self.estoque -= quantidade

        #informa ao usuário a transação
        print(f'você comprou {self.quantidade} ingresso(s)!\nAinda temos {self.estoque} ingressos!')

        #agora, vamos liberar a entrada de um outro thread
        self.lock.release()

#inicializamos o objeto
ingressos = Ingresso2(estoque=20)
if __name__ == '__main': #para inicializar, basta colocar o dunder depois do name 
    for i in range(20):
        t = Thread(target=ingressos.comprar, args=(5,)).start()
    '''
    e como pode ver, funcionou normalmente! quando você usa o lock acquire, é como se você estivesse entregando a chave de um
    banheiro para uma thread e após sua entrada, a thread tranca essa porta. Com isso, as demais threads que quiserem entrar
    no banheiro, vai precisar esperar a thread liberar a saída e sair. Esse é o release, que destranca a porta desse tal banheiro.
    Mas não precisa necessariamente uma thread sair do banheiro, pois se vocẽ usa o release e a thread ainda está lá dentro,
    uma outra thread entra e vira uma espécie de banheiro compartilhado.  
    
    '''
    #também é possível replicar esse exemplo com context manager, que é a forma mais fácil (with)

#######################################################################################################
cls()
# # PyPDF2 para manipular arquivos PDF (Instalação)
# PyPDF2 é uma biblioteca de manipulação de arquivos PDF feita em Python puro,
# gratuita e de código aberto. Ela é capaz de ler, manipular, escrever e unir
# dados de arquivos PDF, assim como adicionar anotações, transformar páginas,
# extrair texto e imagens, manipular metadados, e mais.
# A documentação contém todas as informações necessárias para usar PyPDF2.
# Link: https://pypdf2.readthedocs.io/en/3.0.0/
# Ative seu ambiente virtual
# pip install pypdf2

#no caso desse módulo, os pdf's não costumam seguir o padrão convencional, logo,
# a aplicação de códigos para manipulação de um pdf pode ser diferente para manipulação
# de outro. Portanto, é importante fazer a leitura da documentação.

#Agora, através de Path, vou apontar para o diretório que o pdf se encontra.
#de início, vou organizar o diretório. Farei isso através de Path 
from pathlib import Path
RAIZ = Path(__file__).parent   #raiz do módulo
#Agora, vamos criar uma pasta para a aula.
PASTA_AULA = RAIZ / 'aula329' #definindo caminho
PASTA_AULA.mkdir(exist_ok=True) #efetivando criação

#vamos criar duas pastas dentro da pasta da aula. Uma vai armazenar o nosso pdf e a outra vai salvar
# as cópias ou divisões dele.
PDF_ORIGINAL_GIT = PASTA_AULA / 'pdf_original'
PASTA_NOVA = PASTA_AULA / 'arquivo_novo'
#aqui efetivamos a criação da pasta
for __ in (PDF_ORIGINAL_GIT, PASTA_NOVA):
    __.mkdir(exist_ok=True)
#agora, manualmente, vou pegar um pdf e colocar dentro da pasta pdf_original

#Após isso, vamos definir o diretório do arquivo pdf que será analisado
PDF_CAMINHO = PDF_ORIGINAL_GIT / 'git.pdf'

#vamos fazer a importação da classe PdfReader, para leitura
from PyPDF2 import PdfReader

reader = PdfReader(stream=PDF_CAMINHO) #inicializar o objeto

#e agora, vamos pegar todas as páginas do pdf
paginas = reader.pages 

#se você quiser saber o número de páginas, basta fazer uso do len no objeto pages
print(len(paginas)) #-> nesse caso, temos 3 páginas.

#cada página é um iterável, ou seja, eu consigo acessar individualmente através de um iterador ou através de índices
print(paginas[0]) #aqui, eu acesso a página1 através de índice 0 (que é o primeiro). O retorno disso será as informações gerais de um pdf

#se eu quiser acessar somente o texto disso, basta utilizar o método extract_text.
print(paginas[0].extract_text())

#se houver imagens na página 1 do pdf, será retornado todas as informações binárias referentes a ela.
print(paginas[0].images) #como não tem, vai retornar uma lista vazia.

#Também é possível fazer outras coisas (como extração de imagem), por isso é importante consultar a documentação.

##########
#Para escrita, utilizamos a classe pdfWriter
from PyPDF2 import PdfWriter

#Vamos inicializar a classe através de um objeto
writer = PdfWriter()
#agora, vamos pegar a página0 do pdf anterior como referência e passar para o addpage
writer.add_page(paginas[0]) #nesse exato momento, a página 0 do pdf anterior se encontra na ram

#para salvar na memória rom, basta fazer um contex manager e apontar para um lugar na memória.
with open(PASTA_NOVA / 'page0.pdf', 'wb') as fp:#para isso, vamos aproveitar a constante PASTA_NOVA e wb (write bytes)
    writer.write(fp) #agora utilizamos o método write para escrever 

#aqui eu vou exemplificar mais para solidificar o que foi apresentado
with open(PASTA_NOVA / 'pdf_completo.pdf', 'wb') as fp: #aqui apontamos o caminho que o pdf será salvo
    for pagina in reader.pages: #aqui vamos iterar por dentro de cada página
        writer.add_page(pagina) #aqui, vamos adicionar cada página em uma espécie de fila na memória ram
    #e após a fila completar, vamos salvar a fila completa na memória rom, resultando em um pdf completo!
    writer.write(fp)

#Caso você queira fazer a meclagem de um pdf com outro (o que é possível fazer com a classe PdfWriter),
# vamos utilizar a classe PdfMerger

from PyPDF2 import PdfMerger
#inicializando a classe através de um objeto
merger = PdfMerger()

#agora, vou apontar para dois pdf's dentro de uma lista.
pdfs = [
    PASTA_NOVA / 'page0.pdf', #tem que informar exatamente o nome do pdf
    PASTA_NOVA / 'pdf_completo.pdf'
]

#agora, faremos um for, juntamente com um context manager para colocar na memória
with open(PASTA_NOVA / 'pdf_mesclado.pdf', 'wb') as fp:
    #iterando nos dois pdfs
    for pdf in pdfs:
        #adicionando os pdfs na fila individualmente
        merger.append(pdf)
    #efetivando o merge
    merger.write(fp)

##########################################################################
#-----------------------
# Explicação resumida 

#Fifo -> First input First output (primeiro a entrar primeiro a sair) Significa que o primeiro iterável a entrar numa lista,
#será o primeiro a sair. Ou seja, inicialmente, trata-se do item[0]. 
# ex: items = [1,2,3]. Logo, o primeiro a sair seria 1
# Explicação resumida: Lifo trabalha com o primeiro índice da lista, ou seja, conforme o usuário faz modificação no primeiro item,
#os demais itens terão de se realocar e conforme o tamanho da lista, isso pode demandar tempo e recursos. É por isso que o Lifo
#é denominado de "tempo linear", pois depende do tamanho da lista.
# Adicional: Em caso de listas grandes que precisem de lifo, é de extrema importância utilizar o "deque", pois garante maior
#eficiência de recursos (como tempo) para execução. Deque é específica para esse tipo de coisa, podendo adicionar elementos ou 
#remover elementos da lista.


#Lifo -> Last Input First Output (último a entrar primeiro a sair) Significa que o último iterável a entrar numa lista,
# será o primeiro a sair. Inicialmente, trata-se do índice [-1]
# ex:  items = [1,2,3]. Logo, o primeiro a sair será o 3.
# Explicação resumida: Lifo trabalha com o último índice (pop), não modificando o primeiro. Por isso, o tempo dela é considerado constante
#independente do tamanho da lista

#----------------------
# Deque - Trabalhando com LIFO e FIFO
# deque - Double-ended queue
#
# Lifo  e fifo
# pilha e fila


# LIFO (Last In First Out)
# Pilha (stack)
# Significa que o último item a entrar será o primeiro a sair (list)
# Artigo:
# https://www.otaviomiranda.com.br/2020/pilhas-em-python-com-listas-stack/
# Vídeo:
# https://youtu.be/svWVHEihyNI
# Para tirar itens do final: O(1) Tempo constante
# Para tirar itens do início: O(n) Tempo Linear

from collections import deque

lista = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]

# ✅ Legal (LIFO com lista)
#  0  1  2  3  4  5  6  7  8  9
# [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
lista.append(10)
#  0  1  2  3  4  5  6  7  8  9  10
# [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
lista.append(11)
#  0  1  2  3  4  5  6  7  8  9  10, 11
# [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
ultimo_removido = lista.pop()
#  0  1  2  3  4  5  6  7  8  9  10
# [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
print('Último: ', ultimo_removido)
print('Lista:', lista)
#  0  1  2  3  4  5  6  7  8  9  10
# [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
print()


lista = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]

# 🚫 Ruim (FIFO com lista)
#  0  1  2  3  4  5  6  7  8  9
# [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
lista.insert(0, 10)
#   0  1  2  3  4  5  6  7  8  9, 10
# [10, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
lista.insert(0, 11)
#  0   1   2  3  4  5  6  7  8  9, 10 11
# [11, 10, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
primeiro_removido = lista.pop(0)  # 11
#  0   1  2  3  4  5  6  7  8  9, 10
# [10, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
print('Primeiro: ', primeiro_removido)  # 11
print('Lista:', lista)  # [10, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
print()

# FIFO (First In First Out)
# Filas (queue)
# Significa que o primeiro item a entrar será o primeiro a sair (deque)
# Artigo:
# https://www.otaviomiranda.com.br/2020/filas-em-python-com-deque-queue/
# Vídeo:
# https://youtu.be/RHb-8hXs3HE
# Para tirar itens do final: O(1) Tempo constante
# Para tirar itens do início: O(1) Tempo constante

# ✅ Legal (FIFO com deque)

fila_correta: deque[int] = deque()
fila_correta.append(3)  # Adiciona no final
fila_correta.append(4)  # Adiciona no final
fila_correta.append(5)  # Adiciona no final
fila_correta.appendleft(2)  # Adiciona no começo
fila_correta.appendleft(1)  # Adiciona no começo
fila_correta.appendleft(0)  # Adiciona no começo
print(fila_correta)  # deque([0, 1, 2, 3, 4, 5])
fila_correta.pop()  # 5
fila_correta.popleft()  # 0
print(fila_correta)  # deque([1, 2, 3, 4])

#########################################################################################################33
#O módulo openpyxl é uma alternativa para trabalhar com arquivos Excel xlsx, xlsm, xltx e xltm.
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos a grandes
#quantidades de dados.
# Documentação: https://openpyxl.readthedocs.io/en/stable/ 
#para fazer uso do openpyxl, é necessário uma instalação externa. 
# Sendo: "pip install openpyxl" no terminal. De preferência, após ativar venv.

#após isso, podemos importar o openpyxl normalmente.
import openpyxl 

#para ajudar na manipulação de diretórios, a classe Path seria de grande ajuda.
from pathlib import Path
#Definindo o caminho da pasta raiz
ROOT_FOLDER = Path(__file__).parent
#Definindo o caminho da pasta que ficará armazenado o arquivo
CLASS_FOLDER = ROOT_FOLDER / 'aula335, 336 e 337'
#realizando a criação da pasta
CLASS_FOLDER.mkdir(exist_ok=True)
#definindo o caminho completo do arquivo xlsx, especificando o nome do arquivo e formato.
WORKBOOK_PATH = CLASS_FOLDER / 'workbook.xlsx' #pasta de trabalho

#vamos inicializar a classe através de um objeto
workbook = openpyxl.Workbook()
#e agora vamos inicializar a tipagem, para trabalharmos com as planilhas mais facilmente
worksheet = workbook.active

#o método cell tem três parâmetros, sendo respectivamente:
#row -> linha
#culumn -> coluna
#value -> valor que será armazenado.
#Com isso, passamos essa informação. Abaixo, selecionei coluna e linha 1. Após isso, definimi o valor.
worksheet.cell(1, 1, "Nome")
worksheet.cell(1, 2, "Idade")
worksheet.cell(1, 3, "Nota")

#Se eu quiser verificar os nomes das Abas em minha planilha, basta utilizar o atributo sheetnames. ex:
print(workbook.sheetnames)

#Se eu quiser criar uma planilha nova (uma aba), basta utilizar o método create_sheet, onde:
#title -> nome da aba/planilha que eu desejo inserir
#index -> o índice/posição que eu quero que fique aquela aba.
# Essas abas que eu falo é aquelas que ficam no canto inferior esquerdo na planilha
workbook.create_sheet(title='Apagar Depois', index=0) 

#e caso eu queira apagar uma aba, basta utilizar o método remove (se não existir, vai dar KeyError)
workbook.remove(workbook['Apagar Depois'])

#e para realizar a criação da planilha, basta chamar o OBJETO workbook e utilizar o método save.
workbook.save(WORKBOOK_PATH)

####### Para maior utilidade, vou fazer uma outra planilha que gerará dados mais interessantes.
WORKBOOK_PATH1 = CLASS_FOLDER / 'workbook1.xlsx' #aproveito do diretório criado por path anteriormente

#Como deve ter percebido, a manipulação de tabelas é com base em linhas e colunas de uma forma
# bem semelhante a uma matriz, correto? Sabendo disso,vamos criar uma tabela mais útil utilizando matriz.

#criando uma lista de forma estruturada (em produção não será criada com esse tanto de espaço.)
students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

#Para saber a linha e coluna, vamos precisar de enumerate.s
for i, students_row in enumerate(students, start=2): #começa de dois, pois a lista se inicia do 0 e o índice 1 já foi preenchido
    print(f'Por causa de enumerate, será retornado a linha "{i}"\n e a uma sublista, sendo: "{students_row}\n"') 

    #aproveitando do retorno de student_row, vamos coletar a numeração da coluna
    for j, students_col in enumerate(students_row, start=1): #começará de 1, pois em python, a lista começa do índice 0 e a planilha do 1
        print(f'o índice da coluna é: {j}\nE agora, será retornado o valor literal, sendo: {students_col}')
        
        #Por, basta colocar as linhas e colunas no método cell
        #coloca a numeração da linha, da coluna e o retorno final do value, que estará em student_col.
        worksheet.cell(row=i, column=j, value=students_col)
#para salvar, chamo a pasta de trabalho e passo o diretório da nova pasta
workbook.save(WORKBOOK_PATH1)

#### Para deixar ainda mais simples, a classe tem por padrão uma iteração matriz de forma simplificada. Sendo necessário
#apenas denominar uma tabela, como a nossa anterior e passar como argumento para um método.

#Vou exemplificar. Para isso, criarei um novo caminho PATH, reaproveitando dos anteriores
WORKBOOK_PATH2 = CLASS_FOLDER / 'workbook2.xlsx'

#Vou aproveitar do objeto students, que é uma lista dentro de uma lista.

#Faço um for simples na tabela
for student in students:
    #e chamo o método append para inserir individualmente as linhas
    worksheet.append(student)

#para finalizar, basta chamar o método save e apontar o diretório
workbook.save(WORKBOOK_PATH2)
#Só falta verificar como muda o nome de uma aba padrão.

cls()
########## 
# ler dados de uma planilha

#para isso, vamos importar o método load_workbook da library openpyxl
from openpyxl import load_workbook #type:ignore

#aqui vamos reaproveitar a variável anterior para apontar onde a planilha se encontra
WORKBOOK_PATH = CLASS_FOLDER / 'workbook2.xlsx' #pasta de trabalho

#inicializamos o método e apontamos o diretório da planilha
workbook_r = load_workbook(WORKBOOK_PATH)

#inicializamos as alterações através de um objeto e salvamos o estado
worksheet_r = workbook_r.active

#caso a tipagem do worksheet_r seja unknown e você desejar ter auxílio de tips hints, você ṕde tipar chamando a seguinte classe: 
from openpyxl.cell import Cell #type: ignore

#como não há como tipar no for, vamos tipar externamente
row: Cell

#detalhe: você não precisa necessariamente fazer isso para que o código funcione. 

#agora vamos iterar sobre a planilha
for row in worksheet_r.iter_rows():
    for col in row:
        #aqui conseguimos ver na tela o conteúdo que lá dentro tem
        print(col.value, end='\t')
    print()

#e também consigo fazer modificações, nesse caso, incrementando 17 como valor para idade.
worksheet_r['B3'].value = 17

#por fim, basta salvar e as modificações se aplicarão.
workbook_r.save(WORKBOOK_PATH)

############################################################################################




